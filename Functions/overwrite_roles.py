from __future__ import annotations

from copy import copy
import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _norm_cell(x) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    if s.lower() in {"nan", "none"}:
        return ""
    return s


def _make_row_key(emp_id: str, mgr_id: str, title: str, team: str) -> str:
    """
    Priority:
      1) EMP::<emp_id>
      2) MGR_TITLE_TEAM::<mgr_id>|<title>|<team>
      3) TITLE_TEAM::<title>|<team>
      else ""
    """
    emp_id = _norm_cell(emp_id)
    mgr_id = _norm_cell(mgr_id)
    title = _norm_cell(title)
    team = _norm_cell(team)

    if emp_id:
        return f"EMP::{emp_id}"
    if mgr_id and title and team:
        return f"MGR_TITLE_TEAM::{mgr_id}|{title}|{team}"
    if title and team:
        return f"TITLE_TEAM::{title}|{team}"
    return ""


def _copy_cell_style(src_cell, dst_cell) -> None:
    dst_cell._style = copy(src_cell._style)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)
    dst_cell.alignment = copy(src_cell.alignment)


def _get_or_create_col_near(
    *,
    col_name: str,
    anchor_idx: int,
    header_row: int,
    ws,
    header: dict,
    style_header_from_idx: int,
    style_body_from_idx: int,
) -> int:
    col_name = str(col_name).strip()
    if not col_name:
        raise ValueError("col_name cannot be empty")

    if col_name in header:
        return header[col_name]

    col = anchor_idx + 1
    while True:
        v = ws.cell(row=header_row, column=col).value
        if v is None or str(v).strip() == "":
            break
        col += 1

    dst_hdr = ws.cell(row=header_row, column=col)
    dst_hdr.value = col_name

    src_hdr = ws.cell(row=header_row, column=style_header_from_idx)
    _copy_cell_style(src_hdr, dst_hdr)

    src_letter = get_column_letter(style_header_from_idx)
    dst_letter = get_column_letter(col)
    ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width

    for r in range(header_row + 1, ws.max_row + 1):
        src_cell = ws.cell(row=r, column=style_body_from_idx)
        dst_cell = ws.cell(row=r, column=col)
        _copy_cell_style(src_cell, dst_cell)

    header[col_name] = col
    return col


def write_mapping_to_copy(
    input_xlsm_path: str,
    output_folder: str,
    df_final: pd.DataFrame,
    *,
    sheet_name: str = "Employee census",
    header_row: int = 13,
    # --- input file columns (Excel) ---
    emp_id_col: str = "Employee ID",
    sup_id_col: str = "Manager ID",
    title_col: str = "Job-Profile",
    team_col: str = "Team",
    gtm_col: str = "GTM roles mapped",  # kept for keying + styling; NOT overwritten
    # --- df_final columns (working file) ---
    final_l1_col: str = "Final_Mapped_L1",
    l2_write_col: str = "Mapped_L2_to_write",
    l1_conf_col: str = "L1_Confidence",
    output_suffix: str = "_with_GTM_mapping",
    # --- new output columns to create/write (Excel) ---
    out_l1_header: str = "Mapped_L1",
    out_l1_conf_header: str = "Mapped_L1_Confidence",
    out_l2_header: str = "Mapped_L2",
):
    """
    Writes mapping results into NEW columns in a copied .xlsm workbook, preserving macros,
    formatting, and row alignment.

    - DOES NOT overwrite existing Team or GTM roles mapped columns.
    - Creates/uses NEW columns:
        - out_l1_header populated from df_final[final_l1_col]
        - out_l1_conf_header populated from df_final[l1_conf_col]
        - out_l2_header populated from df_final[l2_write_col]
    - Places new columns near existing table columns (no blank gap).
    - Copies header + body formatting from reference columns for consistent styling.
    - Output path/location logic unchanged.
    """

    # ------------------------------------------------------------------
    # 1) Create a copy of the input file (OUTPUT LOCATION UNCHANGED)
    # ------------------------------------------------------------------
    output_folder = output_folder or "."
    os.makedirs(output_folder, exist_ok=True)

    base_name = os.path.basename(input_xlsm_path)
    base_root, ext = os.path.splitext(base_name)
    output_path = os.path.join(output_folder, f"{base_root}{output_suffix}{ext}")

    shutil.copy2(input_xlsm_path, output_path)

    # ------------------------------------------------------------------
    # 2) Build lookup maps from df_final using robust keys
    # ------------------------------------------------------------------
    required_df_cols = [title_col, team_col, final_l1_col, l2_write_col, l1_conf_col]
    missing_df = [c for c in required_df_cols if c not in df_final.columns]
    if missing_df:
        raise KeyError(f"df_final missing required columns: {missing_df}")

    df_has_emp = emp_id_col in df_final.columns
    df_has_mgr = sup_id_col in df_final.columns

    tmp = df_final.copy()

    tmp["_emp_id_"] = tmp[emp_id_col].apply(_norm_cell) if df_has_emp else ""
    tmp["_mgr_id_"] = tmp[sup_id_col].apply(_norm_cell) if df_has_mgr else ""
    tmp["_title_"] = tmp[title_col].apply(_norm_cell)
    tmp["_team_"] = tmp[team_col].apply(_norm_cell)

    tmp["_key_"] = tmp.apply(
        lambda r: _make_row_key(r["_emp_id_"], r["_mgr_id_"], r["_title_"], r["_team_"]),
        axis=1,
    )

    tmp = tmp[tmp["_key_"].astype(bool)].copy()
    tmp = tmp.drop_duplicates(subset=["_key_"], keep="last")

    l1_map = dict(zip(tmp["_key_"], tmp[final_l1_col].astype(str)))
    l2_map = dict(zip(tmp["_key_"], tmp[l2_write_col].astype(str)))

    # Confidence: keep blank if NaN
    conf_series = pd.to_numeric(tmp[l1_conf_col], errors="coerce")
    l1c_map = dict(zip(tmp["_key_"], conf_series))

    # ------------------------------------------------------------------
    # 3) Load the copied workbook (keep macros)
    # ------------------------------------------------------------------
    wb = load_workbook(output_path, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    # ------------------------------------------------------------------
    # 4) Locate column indices from header row + create new output columns
    # ------------------------------------------------------------------
    header = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is not None:
            header[str(v).strip()] = col

    missing_excel = [c for c in [title_col, team_col, gtm_col] if c not in header]
    if missing_excel:
        raise KeyError(f"Missing columns in header row {header_row}: {missing_excel}")

    title_idx = header[title_col]
    team_idx = header[team_col]
    gtm_idx = header[gtm_col]
    emp_idx = header.get(emp_id_col)
    mgr_idx = header.get(sup_id_col)

    # Create output columns near the existing table columns
    mapped_l1_idx = _get_or_create_col_near(
        col_name=out_l1_header,
        anchor_idx=team_idx,
        header_row=header_row,
        ws=ws,
        header=header,
        style_header_from_idx=team_idx,
        style_body_from_idx=team_idx,
    )

    # Put confidence right next to Mapped_L1 (so they appear together)
    mapped_l1_conf_idx = _get_or_create_col_near(
        col_name=out_l1_conf_header,
        anchor_idx=mapped_l1_idx,
        header_row=header_row,
        ws=ws,
        header=header,
        style_header_from_idx=team_idx,
        style_body_from_idx=team_idx,
    )

    mapped_l2_idx = _get_or_create_col_near(
        col_name=out_l2_header,
        anchor_idx=gtm_idx,
        header_row=header_row,
        ws=ws,
        header=header,
        style_header_from_idx=gtm_idx,
        style_body_from_idx=gtm_idx,
    )

    # ------------------------------------------------------------------
    # 5) Update rows (write into new columns only)
    # ------------------------------------------------------------------
    updated_rows = 0
    skipped_rows_no_key = 0

    for r in range(header_row + 1, ws.max_row + 1):
        emp_val = ws.cell(row=r, column=emp_idx).value if emp_idx else None
        mgr_val = ws.cell(row=r, column=mgr_idx).value if mgr_idx else None
        title_val = ws.cell(row=r, column=title_idx).value
        team_val = ws.cell(row=r, column=team_idx).value

        key = _make_row_key(
            _norm_cell(emp_val),
            _norm_cell(mgr_val),
            _norm_cell(title_val),
            _norm_cell(team_val),
        )

        if not key:
            skipped_rows_no_key += 1
            continue

        wrote = False

        if key in l1_map:
            ws.cell(row=r, column=mapped_l1_idx).value = l1_map[key]
            wrote = True

        if key in l1c_map:
            v = l1c_map[key]
            # write blank if NaN
            if v is None or (isinstance(v, float) and pd.isna(v)):
                ws.cell(row=r, column=mapped_l1_conf_idx).value = None
            else:
                ws.cell(row=r, column=mapped_l1_conf_idx).value = float(v)
            wrote = True

        if key in l2_map:
            ws.cell(row=r, column=mapped_l2_idx).value = l2_map[key]
            wrote = True

        if wrote:
            updated_rows += 1

    # ------------------------------------------------------------------
    # 6) Save
    # ------------------------------------------------------------------
    wb.save(output_path)

    return output_path, updated_rows, skipped_rows_no_key
